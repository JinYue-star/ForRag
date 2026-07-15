#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""课堂练习：题库 CSV/XLSX 解析、模板、题库导出、学生提问入库。"""

from __future__ import annotations

import csv
import io
import re
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import chroma_store
import exercise_store
import kb_store

from rag_api import settings
from rag_api.http_common import safe_filename
from rag_api.qa_llm import (
    build_quiz_public,
    coerce_quiz_index,
    generate_quiz_bundle_from_student_questions,
    quiz_generation_fail_detail,
)

BANK_COLUMNS = [
    "type",
    "question",
    "option1",
    "option2",
    "option3",
    "option4",
    "option5",
    "option6",
    "correct",
]

STUDENT_QUESTIONS_CATEGORY = "Student questions"

TEMPLATE_SAMPLE_ROWS = [
    {
        "type": "tf",
        "question": "HTTP is a transport-layer protocol.",
        "option1": "True",
        "option2": "False",
        "option3": "",
        "option4": "",
        "option5": "",
        "option6": "",
        "correct": "False",
    },
    {
        "type": "single",
        "question": "Which layer does IP belong to?",
        "option1": "Application",
        "option2": "Transport",
        "option3": "Network",
        "option4": "Link",
        "option5": "",
        "option6": "",
        "correct": "C",
    },
    {
        "type": "multi",
        "question": "Select reliable transport features.",
        "option1": "ACK",
        "option2": "Best-effort delivery only",
        "option3": "Retransmission",
        "option4": "Unordered datagrams only",
        "option5": "",
        "option6": "",
        "correct": "A|C",
    },
]


def _letter_to_index(token: str) -> Optional[int]:
    t = token.strip().upper()
    if len(t) == 1 and "A" <= t <= "Z":
        return ord(t) - 65
    return coerce_quiz_index(t)


def _parse_correct_tokens(raw: str) -> list[str]:
    s = (raw or "").strip()
    if not s:
        return []
    parts = re.split(r"[|,;]+", s)
    return [p.strip() for p in parts if p.strip()]


def _options_from_row(row: dict[str, str]) -> list[str]:
    opts: list[str] = []
    for i in range(1, 7):
        v = (row.get(f"option{i}") or "").strip()
        if v:
            opts.append(v)
    return opts


def _normalize_header(name: str) -> str:
    return re.sub(r"\s+", "", (name or "").strip().lower())


HEADER_ALIASES = {
    "type": "type",
    "question": "question",
    "option1": "option1",
    "option2": "option2",
    "option3": "option3",
    "option4": "option4",
    "option5": "option5",
    "option6": "option6",
    "correct": "correct",
    "answer": "correct",
    "正确答案": "correct",
    "题型": "type",
    "题目": "question",
}


def _map_headers(raw_headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for i, h in enumerate(raw_headers):
        key = HEADER_ALIASES.get(_normalize_header(h))
        if key and key not in mapping:
            mapping[key] = i
    return mapping


def _row_dict(headers_map: dict[str, int], cells: list[str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for key, idx in headers_map.items():
        out[key] = cells[idx] if idx < len(cells) else ""
    return out


def parse_bank_rows(rows: list[dict[str, str]]) -> tuple[list[dict[str, Any]], list[str]]:
    """将表格行解析为内部 quiz items；返回 (items, errors)。任一错误则 items 为空。"""
    items: list[dict[str, Any]] = []
    errors: list[str] = []
    for line_no, row in enumerate(rows, start=2):  # 1-based header + data
        t = (row.get("type") or "").strip().lower()
        q = (row.get("question") or "").strip()
        if not t and not q and not any((row.get(f"option{i}") or "").strip() for i in range(1, 7)):
            continue
        if t not in ("tf", "single", "multi"):
            errors.append(f"Row {line_no}: type must be tf, single, or multi")
            continue
        if not q:
            errors.append(f"Row {line_no}: question is required")
            continue
        opts = _options_from_row(row)
        correct_raw = row.get("correct") or ""
        tokens = _parse_correct_tokens(correct_raw)
        if t == "tf":
            if len(opts) == 0:
                opts = ["True", "False"]
            if len(opts) != 2:
                errors.append(f"Row {line_no}: tf needs exactly 2 options (or leave blank for True/False)")
                continue
            a0, a1 = opts[0].casefold(), opts[1].casefold()
            if {a0, a1} != {"true", "false"}:
                # allow arbitrary labels but still store as given; correct must resolve
                pass
            if not tokens:
                errors.append(f"Row {line_no}: correct is required")
                continue
            tok = tokens[0]
            ci: Optional[int] = None
            if tok.casefold() in ("true", "false"):
                for i, o in enumerate(opts):
                    if o.casefold() == tok.casefold():
                        ci = i
                        break
            else:
                ci = _letter_to_index(tok)
            if ci is None or ci not in (0, 1):
                errors.append(f"Row {line_no}: invalid tf correct value")
                continue
            items.append(
                {
                    "type": "tf",
                    "question": q,
                    "options": ["True", "False"] if {opts[0].casefold(), opts[1].casefold()} == {"true", "false"} else opts,
                    "correct_index": ci if {opts[0].casefold(), opts[1].casefold()} != {"true", "false"} else (
                        0 if opts[ci].casefold() == "true" or (tok.casefold() == "true") else 1
                    ),
                }
            )
            # Normalize classic True/False
            if {opts[0].casefold(), opts[1].casefold()} == {"true", "false"}:
                word = tok.casefold() if tok.casefold() in ("true", "false") else opts[ci].casefold()
                items[-1]["options"] = ["True", "False"]
                items[-1]["correct_index"] = 0 if word == "true" else 1
        elif t == "single":
            if len(opts) < 2:
                errors.append(f"Row {line_no}: single needs at least 2 options")
                continue
            if not tokens:
                errors.append(f"Row {line_no}: correct is required")
                continue
            ci = _letter_to_index(tokens[0])
            if ci is None or ci < 0 or ci >= len(opts):
                errors.append(f"Row {line_no}: invalid single correct value")
                continue
            items.append({"type": "single", "question": q, "options": opts, "correct_index": ci})
        else:  # multi
            if len(opts) < 2:
                errors.append(f"Row {line_no}: multi needs at least 2 options")
                continue
            if len(tokens) < 1:
                errors.append(f"Row {line_no}: correct is required for multi")
                continue
            cis: list[int] = []
            for tok in tokens:
                j = _letter_to_index(tok)
                if j is None or j < 0 or j >= len(opts):
                    errors.append(f"Row {line_no}: invalid multi correct token '{tok}'")
                    cis = []
                    break
                cis.append(j)
            if not cis:
                continue
            cis = sorted(set(cis))
            if len(cis) < 1:
                errors.append(f"Row {line_no}: multi needs at least one correct option")
                continue
            # Align with LLM normalizer which prefers >=2 for multi; allow 1 for teacher banks
            items.append({"type": "multi", "question": q, "options": opts, "correct_indices": cis})
    if errors:
        return [], errors
    if not items:
        return [], ["No question rows found"]
    return items, []


def parse_bank_bytes(content: bytes, filename: str) -> tuple[list[dict[str, Any]], list[str]]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        try:
            header = next(reader)
        except StopIteration:
            return [], ["Empty CSV"]
        hmap = _map_headers(header)
        if "type" not in hmap or "question" not in hmap or "correct" not in hmap:
            return [], ["CSV must include columns: type, question, correct"]
        rows: list[dict[str, str]] = []
        for cells in reader:
            rows.append(_row_dict(hmap, [str(c) if c is not None else "" for c in cells]))
        return parse_bank_rows(rows)

    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            return [], ["openpyxl is required for Excel import"]
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            header_cells = next(it)
        except StopIteration:
            return [], ["Empty Excel sheet"]
        header = ["" if c is None else str(c) for c in header_cells]
        hmap = _map_headers(header)
        if "type" not in hmap or "question" not in hmap or "correct" not in hmap:
            return [], ["Excel must include columns: type, question, correct"]
        rows = []
        for cells in it:
            vals = ["" if c is None else str(c) for c in cells]
            rows.append(_row_dict(hmap, vals))
        return parse_bank_rows(rows)

    return [], ["Only .csv or .xlsx files are supported"]


def items_to_bank_rows(items: list[dict[str, Any]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for it in items:
        t = str(it.get("type") or "").lower()
        opts = [str(x) for x in (it.get("options") or [])]
        row = {c: "" for c in BANK_COLUMNS}
        row["type"] = t
        row["question"] = str(it.get("question") or "")
        for i, o in enumerate(opts[:6], start=1):
            row[f"option{i}"] = o
        if t == "multi":
            cis = it.get("correct_indices") or []
            letters = []
            for j in cis:
                jj = coerce_quiz_index(j)
                if jj is not None:
                    letters.append(chr(65 + jj))
            row["correct"] = "|".join(letters)
        else:
            ci = coerce_quiz_index(it.get("correct_index"))
            if ci is not None:
                if t == "tf" and opts and 0 <= ci < len(opts):
                    row["correct"] = opts[ci]
                else:
                    row["correct"] = chr(65 + ci)
        rows.append(row)
    return rows


def bank_to_csv(items: list[dict[str, Any]]) -> bytes:
    sio = io.StringIO()
    w = csv.DictWriter(sio, fieldnames=BANK_COLUMNS)
    w.writeheader()
    for row in items_to_bank_rows(items):
        w.writerow(row)
    return sio.getvalue().encode("utf-8-sig")


def bank_to_xlsx(items: list[dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append(BANK_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in items_to_bank_rows(items):
        ws.append([row.get(c, "") for c in BANK_COLUMNS])
    for col_idx, name in enumerate(BANK_COLUMNS, start=1):
        width = 48 if name == "question" else 14
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def template_csv_bytes() -> bytes:
    sio = io.StringIO()
    w = csv.DictWriter(sio, fieldnames=BANK_COLUMNS)
    w.writeheader()
    for r in TEMPLATE_SAMPLE_ROWS:
        w.writerow(r)
    return sio.getvalue().encode("utf-8-sig")


def template_xlsx_bytes() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append(BANK_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in TEMPLATE_SAMPLE_ROWS:
        ws.append([r.get(c, "") for c in BANK_COLUMNS])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def create_exercise_from_items(
    items: list[dict[str, Any]],
    *,
    title: str,
    created_by: Optional[str],
    source_filename: Optional[str] = None,
    source_note_id: Optional[str] = None,
    status: str = exercise_store.STATUS_PUBLISHED,
) -> dict[str, Any]:
    quiz_id = uuid.uuid4().hex
    now = time.time()
    payload = {"items": items, "meta": {"source": "class_exercise", "title": title}}
    chroma_store.quiz_insert(quiz_id, None, payload, now)
    row = exercise_store.exercise_insert(
        settings.DATA_DIR,
        settings.KB_ID,
        title=title,
        quiz_id=quiz_id,
        item_count=len(items),
        status=status,
        source_filename=source_filename,
        source_note_id=source_note_id,
        created_by=created_by,
    )
    return row


def delete_exercise(exercise_id: str) -> bool:
    row = exercise_store.exercise_delete(settings.DATA_DIR, settings.KB_ID, exercise_id)
    if not row:
        return False
    qid = str(row.get("quiz_id") or "")
    if qid:
        chroma_store.quiz_delete(qid)
        chroma_store.quiz_answers_delete_for_quiz(qid)
    return True


def ensure_student_questions_category(owner_id: Optional[str] = None) -> dict[str, Any]:
    cats = kb_store.categories_list(settings.DATA_DIR, settings.KB_ID)
    for c in cats:
        if str(c.get("name") or "").strip().casefold() == STUDENT_QUESTIONS_CATEGORY.casefold():
            return c
    return kb_store.category_insert(
        settings.DATA_DIR,
        settings.KB_ID,
        STUDENT_QUESTIONS_CATEGORY,
        owner_id=owner_id,
        sort_order=0,
    )


def save_questions_to_kb(
    question_rows: list[dict[str, Any]],
    *,
    fmt: str,
    owner_id: Optional[str] = None,
) -> dict[str, Any]:
    if not question_rows:
        raise ValueError("no_questions")
    cat = ensure_student_questions_category(owner_id)
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    title = f"Student questions — {stamp}"
    lines = [f"# {title}", "", f"Exported {len(question_rows)} student question(s).", ""]
    for i, row in enumerate(question_rows, start=1):
        q = str(row.get("question") or "").strip()
        who = str(row.get("student_name") or row.get("username") or "").strip()
        when = str(row.get("time") or "").strip()
        meta = " · ".join(x for x in (who, when) if x)
        lines.append(f"{i}. {q}" + (f"  \n   _{meta}_" if meta else ""))
    body = "\n".join(lines) + "\n"
    note = kb_store.note_insert(
        settings.DATA_DIR,
        settings.KB_ID,
        str(cat["id"]),
        title,
        body,
        owner_id=owner_id,
    )
    # Attach export file (questions module only)
    from rag_api import export_service

    data = {"questions": question_rows, "answers": [], "quiz": []}
    modules = ["questions"]
    fmt_l = (fmt or "xlsx").lower()
    if fmt_l == "csv":
        payload = export_service.to_csv(data, modules)
        fname = f"student-questions-{datetime.now().strftime('%Y%m%d-%H%M%S')}.csv"
        mime = "text/csv"
    else:
        payload = export_service.to_xlsx(data, modules)
        fname = f"student-questions-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    attach_id = uuid.uuid4().hex
    kb_dir = settings.kb_files_dir()
    kb_dir.mkdir(parents=True, exist_ok=True)
    used = {p.name for p in kb_dir.iterdir() if p.is_file()}
    safe_name = safe_filename(fname, used)
    disk_name = f"{attach_id}_{safe_name}"
    stored_rel = f"{settings.KB_ROOT_REL}/files/{disk_name}"
    dest = settings.UPLOAD_DIR / stored_rel
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(payload)
    kb_store.note_file_insert(
        settings.DATA_DIR,
        settings.KB_ID,
        str(note["id"]),
        fname,
        stored_rel,
        len(payload),
        mime,
    )
    return {
        "category_id": cat["id"],
        "note_id": note["id"],
        "title": title,
        "question_count": len(question_rows),
    }


def extract_questions_from_note_body(body: str) -> list[str]:
    texts: list[str] = []
    for line in (body or "").splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        m = re.match(r"^\d+\.\s+(.*)$", s)
        if m:
            q = m.group(1).strip()
            # strip trailing italic meta on same line
            q = re.sub(r"\s*_.*?_\s*$", "", q).strip()
            if q:
                texts.append(q)
            continue
        if s.startswith("- ") or s.startswith("* "):
            texts.append(s[2:].strip())
    return texts


def generate_and_optionally_publish(
    note_id: str,
    *,
    n: int,
    title: Optional[str],
    publish: bool,
    created_by: Optional[str],
) -> dict[str, Any]:
    note = kb_store.note_get(settings.DATA_DIR, settings.KB_ID, note_id)
    if not note:
        raise ValueError("note_not_found")
    questions = extract_questions_from_note_body(str(note.get("body_markdown") or ""))
    if not questions:
        raise ValueError("no_questions_in_note")
    total = max(1, min(int(n or 5), 20))
    bundle, code = generate_quiz_bundle_from_student_questions(questions, total)
    if not bundle:
        raise RuntimeError(quiz_generation_fail_detail(code))
    items = bundle["items"]
    ex_title = (title or "").strip() or f"From: {note.get('title') or 'Student questions'}"
    result: dict[str, Any] = {
        "title": ex_title,
        "item_count": len(items),
        "items_public": build_quiz_public("preview", items).model_dump(),
        "items": items,
        "exercise": None,
    }
    if publish:
        row = create_exercise_from_items(
            items,
            title=ex_title,
            created_by=created_by,
            source_note_id=note_id,
            status=exercise_store.STATUS_PUBLISHED,
        )
        result["exercise"] = row
        result["items_public"] = build_quiz_public(str(row["quiz_id"]), items).model_dump()
    return result
