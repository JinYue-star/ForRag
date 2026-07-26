#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""教师导出：汇总学生提问与测验作答，序列化为 CSV / Excel。

数据来源：
- 会话（含 owner 归因）+ 用户消息（仅提问文本；不导出回答/完整对话）
- 测验批次（题干/选项/标准答案）+ 学生作答与判分

说明：`include_answers` / 回答摘要模块已停用；路由层固定传 False。
"""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from typing import Any, Optional

import chroma_store

from rag_api.qa_llm import coerce_quiz_index

ANSWER_SUMMARY_MAX = 500

QUESTION_COLUMNS = ["time", "student_name", "student_no", "username", "session_id", "question"]
ANSWER_COLUMNS = [
    "time", "student_name", "student_no", "username", "session_id",
    "question", "answer_summary", "grounded",
]
# Quiz sheet: only assessment fields (aligned with class-exercise import encoding).
QUIZ_COLUMNS = [
    "type", "question", "options", "correct_answer", "student_answer", "is_correct",
]

MODULE_COLUMNS = {
    "questions": QUESTION_COLUMNS,
    "answers": ANSWER_COLUMNS,
    "quiz": QUIZ_COLUMNS,
}
MODULE_SHEET = {"questions": "Questions", "answers": "Answer summaries", "quiz": "Quiz"}


def _fmt_time(ts: float) -> str:
    if not ts:
        return ""
    try:
        return datetime.fromtimestamp(float(ts), tz=timezone.utc).astimezone().strftime(
            "%Y-%m-%d %H:%M:%S"
        )
    except (ValueError, OSError, OverflowError):
        return ""


def _owner_fields(owner: Optional[dict[str, Any]]) -> dict[str, str]:
    o = owner or {}
    return {
        "student_name": str(o.get("display_name") or ""),
        "student_no": str(o.get("student_no") or ""),
        "username": str(o.get("username") or ""),
    }


def _in_range(ts: float, start: Optional[float], end: Optional[float]) -> bool:
    if start is not None and ts < start:
        return False
    if end is not None and ts > end:
        return False
    return True


def _options_cell(item: dict[str, Any]) -> str:
    """与导入题库 option1|option2|… 同一顺序与分隔符。"""
    opts = [str(x).strip() for x in (item.get("options") or []) if str(x).strip()]
    return "|".join(opts[:6])


def _correct_answer_from_item(item: dict[str, Any]) -> str:
    """与课堂练习导入列 `correct` 一致：tf 用 True/False；单选 A/B/…；多选 A|C。"""
    t = str(item.get("type") or "").lower().strip()
    opts = [str(x) for x in (item.get("options") or [])]
    if t == "multi":
        letters: list[str] = []
        for j in item.get("correct_indices") or []:
            jj = coerce_quiz_index(j)
            if jj is not None and 0 <= jj < len(opts):
                letters.append(chr(65 + jj))
        return "|".join(letters)
    ci = coerce_quiz_index(item.get("correct_index"))
    if ci is None or not (0 <= ci < len(opts)):
        return ""
    if t == "tf":
        return opts[ci]
    return chr(65 + ci)


def _student_answer_like_import(item: dict[str, Any], raw: str) -> str:
    """把学生作答规范成与 `correct` / `correct_answer` 相同的字母或 True/False 形式。"""
    t = str(item.get("type") or "").lower().strip()
    opts = [str(x) for x in (item.get("options") or [])]
    s = (raw or "").strip()
    if not s:
        return ""

    def token_to_letter(tok: str) -> str:
        tok = tok.strip()
        if not tok:
            return ""
        if len(tok) == 1 and tok.isalpha():
            return tok.upper()
        idx = coerce_quiz_index(tok)
        if idx is not None and 0 <= idx < len(opts):
            if t == "tf":
                return opts[idx]
            return chr(65 + idx)
        for i, o in enumerate(opts):
            if o == tok or o.casefold() == tok.casefold():
                if t == "tf":
                    return o
                return chr(65 + i)
        return tok

    if t == "multi":
        parts = [p for p in re.split(r"[,|;]+", s) if p.strip()]
        letters = [token_to_letter(p) for p in parts]
        letters = [x for x in letters if x]
        # Keep A|C order by letter
        uniq = sorted(set(letters), key=lambda x: (len(x) != 1, x))
        return "|".join(uniq)
    return token_to_letter(s)


def gather(
    start: Optional[float],
    end: Optional[float],
    course_ids: Optional[list[str]],
    include_questions: bool,
    include_answers: bool,
    include_quiz: bool,
) -> dict[str, list[dict[str, Any]]]:
    """按筛选条件汇总各模块行数据。course_ids 预留多课程（当前单课程忽略）。"""
    sessions = {s["session_id"]: s for s in chroma_store.sessions_list_all()}
    owners = {sid: s.get("owner") for sid, s in sessions.items()}

    data: dict[str, list[dict[str, Any]]] = {"questions": [], "answers": [], "quiz": []}

    if include_questions or include_answers:
        for sid in sessions:
            owner = _owner_fields(owners.get(sid))
            try:
                msgs = chroma_store.messages_list(sid)
            except Exception:
                continue
            for i, m in enumerate(msgs):
                if m.get("role") != "user":
                    continue
                ts = float(m.get("created_at") or 0)
                if not _in_range(ts, start, end):
                    continue
                question = str(m.get("content") or "").strip()
                if include_questions:
                    row = {"time": _fmt_time(ts), "session_id": sid, "question": question}
                    row.update(owner)
                    data["questions"].append(row)
                if include_answers:
                    ans = ""
                    grounded = ""
                    for j in range(i + 1, len(msgs)):
                        if msgs[j].get("role") == "assistant":
                            ans = str(msgs[j].get("content") or "").strip()
                            extra = msgs[j].get("extra") or {}
                            grounded = "yes" if extra.get("kb_relevant") else "no"
                            break
                    row = {
                        "time": _fmt_time(ts),
                        "session_id": sid,
                        "question": question,
                        "answer_summary": ans[:ANSWER_SUMMARY_MAX],
                        "grounded": grounded,
                    }
                    row.update(owner)
                    data["answers"].append(row)

    if include_quiz:
        # Multi-student submissions: one export row set per (quiz, submission)
        submissions_by_quiz: dict[str, list[dict[str, Any]]] = {}
        for sub in chroma_store.quiz_answers_list():
            qid = str(sub.get("quiz_id") or "")
            if not qid:
                continue
            submissions_by_quiz.setdefault(qid, []).append(sub)

        for quiz in chroma_store.quiz_list_all():
            ts = float(quiz.get("created_at") or 0)
            if not _in_range(ts, start, end):
                continue
            sid = quiz.get("session_id") or ""
            qid = quiz.get("quiz_id")
            items = (quiz.get("payload") or {}).get("items") or []
            subs = submissions_by_quiz.get(str(qid) or "", [])
            if not subs:
                # No submissions yet — still export question bank rows with empty student fields
                for it in items:
                    data["quiz"].append(
                        {
                            "type": str(it.get("type") or ""),
                            "question": str(it.get("question") or ""),
                            "options": _options_cell(it),
                            "correct_answer": _correct_answer_from_item(it),
                            "student_answer": "",
                            "is_correct": "",
                        }
                    )
                continue

            for submission in subs:
                grade_items = (submission.get("grade") or {}).get("items") or []
                raw_answers = submission.get("answers") or []
                for idx, it in enumerate(items):
                    g = grade_items[idx] if idx < len(grade_items) else {}
                    correct = _correct_answer_from_item(it)
                    raw_student = ""
                    if g:
                        raw_student = str(g.get("user_answer") or "")
                    elif idx < len(raw_answers):
                        raw_student = str(raw_answers[idx] or "")
                    student_answer = _student_answer_like_import(it, raw_student)
                    is_correct = ""
                    if g:
                        try:
                            is_correct = (
                                "yes"
                                if float(g.get("score", 0)) >= float(g.get("max_score", 0))
                                and float(g.get("max_score", 0)) > 0
                                else "no"
                            )
                        except (TypeError, ValueError):
                            is_correct = ""
                    data["quiz"].append(
                        {
                            "type": str(it.get("type") or ""),
                            "question": str(it.get("question") or ""),
                            "options": _options_cell(it),
                            "correct_answer": correct,
                            "student_answer": student_answer,
                            "is_correct": is_correct,
                        }
                    )

    return data


def selected_modules(include_questions: bool, include_answers: bool, include_quiz: bool) -> list[str]:
    mods: list[str] = []
    if include_questions:
        mods.append("questions")
    if include_answers:
        mods.append("answers")
    if include_quiz:
        mods.append("quiz")
    return mods


def preview(data: dict[str, list[dict[str, Any]]], modules: list[str], limit: int = 50) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for m in modules:
        rows = data.get(m, [])
        out[m] = {
            "columns": MODULE_COLUMNS[m],
            "rows": rows[:limit],
            "total": len(rows),
        }
    return out


def to_csv(data: dict[str, list[dict[str, Any]]], modules: list[str]) -> bytes:
    """输出单个 .csv。多个模块时在同一文件中分段堆叠，段间用标题行分隔。"""
    sio = io.StringIO()
    writer = csv.writer(sio)
    multi = len(modules) > 1
    for mi, m in enumerate(modules):
        cols = MODULE_COLUMNS[m]
        if mi > 0:
            writer.writerow([])
        if multi:
            writer.writerow([f"# {MODULE_SHEET[m]}"])
        writer.writerow(cols)
        for row in data.get(m, []):
            writer.writerow([row.get(c, "") for c in cols])
    return sio.getvalue().encode("utf-8-sig")


def to_xlsx(data: dict[str, list[dict[str, Any]]], modules: list[str]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)
    for m in modules:
        cols = MODULE_COLUMNS[m]
        ws = wb.create_sheet(title=MODULE_SHEET[m][:31])
        ws.append(cols)
        for cell in ws[1]:
            cell.font = header_font
        for row in data.get(m, []):
            ws.append([row.get(c, "") for c in cols])
        # reasonable column widths
        for col_idx, name in enumerate(cols, start=1):
            width = 18 if name not in ("question", "answer_summary", "options") else 48
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    if not modules:
        wb.create_sheet(title="Empty")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
